import os
import urllib.request
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import json
import time
import sys
import base64
import re

CLAN_ID = 105  # ← CHANGE THIS TO YOUR CLAN ID!

API_URL = f"https://playninjarift.com/api/detail_clan_website.php?clan_id={CLAN_ID}"
TARGET_TZ = timezone(timedelta(hours=8))
EXCEL_FILE = f"clan_{CLAN_ID}.xlsx"
HOURLY_CACHE = "_hourly_cache.json"
CACHE_30M = "_30m_cache.json"
CACHE_1H = "_1h_cache.json"
CHANGES_JSON = "_changes.json"
GOAL_TIERS = [
    (100000, "5 Stamina Rolls"),
    (500000, "20 Stamina Rolls"),
    (750000, "Back Item"),
    (1000000, "Weapon"),
    (1600000, "Jutsu"),
]
CHANGES_JSON = "_changes.json"

def fetch_clan():
    req = urllib.request.Request(API_URL, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode())

SEASON_API = "https://playninjarift.com/api/refresh_time_website.php"
RANKING_API = "https://playninjarift.com/api/clan_ranking_website.php"

def fetch_season_info():
    req = urllib.request.Request(SEASON_API, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode())

def fetch_clan_ranking():
    req = urllib.request.Request(RANKING_API, headers={"User-Agent": "clan-snapshot/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode())
    for entry in data:
        if entry["clan_id"] == CLAN_ID:
            return entry
    return {}

def get_previous_sheet_names(wb):
    names = [s.title for s in wb.worksheets]
    names.sort()
    return names

def get_unique_names(members):
    """Returns list of (api_name, display_name) tuples."""
    name_count = {}
    for m in members:
        nm = m["character_name"]
        name_count[nm] = name_count.get(nm, 0) + 1
    counter = {}
    result = []
    for m in members:
        nm = m["character_name"]
        if name_count[nm] > 1:
            counter[nm] = counter.get(nm, 0) + 1
            result.append((nm, f"{nm} (#{counter[nm]})"))
        else:
            result.append((nm, nm))
    return result

def load_prev_from_xlsx(filename, before_date):
    prev_data = []
    prev_timestamp = None
    if not os.path.exists(filename):
        return prev_data, prev_timestamp
    wb = load_workbook(filename)
    prev_names = [s.title for s in wb.worksheets]
    prev_names.sort()
    prev_sheet_name = None
    for n in reversed(prev_names):
        if n < before_date:
            prev_sheet_name = n
            break
    if prev_sheet_name and prev_sheet_name in wb.sheetnames:
        ps = wb[prev_sheet_name]
        raw = ps["A2"].value
        if raw:
            prev_timestamp = raw.replace("Timestamp: ", "")
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[0] and row[1] is not None:
                prev_data.append({"character_name": row[0], "member_reputation": int(row[1])})
    return prev_data, prev_timestamp

def load_hourly_cache():
    if not os.path.exists(HOURLY_CACHE):
        return {}, None
    with open(HOURLY_CACHE, encoding="utf-8") as f:
        c = json.load(f)
    return c.get("members", {}), c.get("timestamp")

def save_hourly_cache(members, unique_names, now):
    cache = {
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "members": {unique_names[i][1]: m["member_reputation"] for i, m in enumerate(members)},
    }
    with open(HOURLY_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_30m_cache():
    if not os.path.exists(CACHE_30M):
        return None
    with open(CACHE_30M, encoding="utf-8") as f:
        return json.load(f)

def save_30m_cache(members, unique_names):
    cache = {
        "timestamp": datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "members": {unique_names[i][1]: m["member_reputation"] for i, m in enumerate(members)},
        "order": [unique_names[i][1] for i, _ in enumerate(members)],
    }
    with open(CACHE_30M, "w", encoding="utf-8") as f:
        json.dump(cache, f)

def load_1h_cache():
    if not os.path.exists(CACHE_1H):
        return None
    with open(CACHE_1H, encoding="utf-8") as f:
        return json.load(f)

def save_1h_cache(member_dict, timestamp):
    with open(CACHE_1H, "w", encoding="utf-8") as f:
        json.dump({"timestamp": timestamp, "members": member_dict}, f)

def compute_rolling_avg_daily_gain(filename, before_date):
    if not os.path.exists(filename):
        return None
    wb = load_workbook(filename)
    names = sorted([s.title for s in wb.worksheets if s.title != "Sheet1" and s.title < before_date])
    gains = []
    for i in range(1, len(names)):
        prev_total = 0
        curr_total = 0
        ps = wb[names[i-1]]
        cs = wb[names[i]]
        for row in ps.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                prev_total += int(row[1])
        for row in cs.iter_rows(min_row=4, max_col=2, values_only=True):
            if row[1] is not None:
                curr_total += int(row[1])
        gains.append(curr_total - prev_total)
    if not gains:
        return None
    return sum(gains) / len(gains)

def compute_changes(members, prev_data):
    prev_names = {m["character_name"] for m in prev_data}
    today_names = {m["character_name"] for m in members}
    left_names = sorted(prev_names - today_names)
    joined_names = sorted(today_names - prev_names)
    return left_names, joined_names

def compute_diff(members, prev_data):
    prev_map = {m["character_name"]: m["member_reputation"] for m in prev_data}
    result = []
    for m in members:
        name = m["character_name"]
        reps = m["member_reputation"]
        if name in prev_map:
            diff = reps - prev_map[name]
            diff_str = f"+{diff}" if diff > 0 else str(diff)
        else:
            diff_str = "N/A"
        result.append((name, reps, diff_str))
    return result

def write_sheet(ws, data, prev_data, now, unique_names):
    uniq = unique_names if unique_names else get_unique_names(data["members"])
    rows = compute_diff(data["members"], prev_data)
    daily_lookup = {name: diff for name, _, diff in rows}
    names = [uniq[i][1] for i in range(len(uniq))]
    reps = [str(m["member_reputation"]) for m in data["members"]]
    diffs = [daily_lookup.get(m["character_name"], "N/A") for m in data["members"]]
    ws.title = now.strftime("%Y-%m-%d")

    max_name = max((len(n) for n in names), default=10)
    max_reps = max((len(r) for r in reps), default=4)
    max_diff = max((len(d) for d in diffs), default=3)

    ws.column_dimensions["A"].width = max(max_name + 5, 10)
    ws.column_dimensions["B"].width = max(max_reps + 3, 8)
    ws.column_dimensions["C"].width = max(max_diff + 3, 14)

    clan_name = data.get("clan_name", "Unknown")
    ws.merge_cells("A1:C1")
    ws["A1"] = f"Clan: {clan_name} ({CLAN_ID})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Timestamp: {now.strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Name", "Total Reps", "Daily Reps (+1d)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (name, reps_val, diff_val) in enumerate(rows, 4):
        ws.cell(row=row_idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=2, value=reps_val).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=diff_val).alignment = Alignment(horizontal="center", vertical="center")

def save_xlsx(data, prev_data, now, uniq):
    sheet_name = now.strftime("%Y-%m-%d")

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    write_sheet(ws, data, prev_data, now, uniq)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(EXCEL_FILE)
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Saved sheet '{sheet_name}' to {EXCEL_FILE}")

def save_seasonal_xlsx(members, season_num):
    filename = f"S{season_num}_ID{CLAN_ID}.xlsx"
    if os.path.exists(filename):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = f"Season {season_num}"
    header = f"[S{season_num}] Total Reps"
    ws["A1"] = "Name"
    ws["B1"] = header
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    max_name = 10
    max_reps = 4
    for idx, m in enumerate(members, 2):
        name = m["character_name"]
        reps = m["member_reputation"]
        ws.cell(row=idx, column=1, value=name).alignment = Alignment(vertical="center")
        ws.cell(row=idx, column=2, value=reps).alignment = Alignment(horizontal="center", vertical="center")
        max_name = max(max_name, len(name))
        max_reps = max(max_reps, len(str(reps)))
    ws.column_dimensions["A"].width = max_name + 5
    ws.column_dimensions["B"].width = max_reps + 3
    wb.save(filename)
    print(f"Saved seasonal snapshot: {filename}")

def compute_season_projection(clan_reputation, avg_daily_gain, season_end_dt, now):
    if avg_daily_gain is None or avg_daily_gain <= 0:
        return None
    projection = clan_reputation
    total_days = 0
    current = now + timedelta(days=1)
    current = current.replace(hour=0, minute=0, second=0, microsecond=0)
    while current <= season_end_dt:
        if current.weekday() in (6, 0):
            projection += avg_daily_gain * 2
        else:
            projection += avg_daily_gain
        total_days += 1
        current += timedelta(days=1)
    days_left = (season_end_dt - now).days
    return {"projection": int(round(projection)), "avg_daily": int(round(avg_daily_gain)), "days_left": days_left}

def compute_goal_info(clan_reputation):
    prev = 0
    prev_name = ""
    for threshold, name in GOAL_TIERS:
        if clan_reputation < threshold:
            progress = clan_reputation - prev
            total = threshold - prev
            pct = (progress / total) * 100 if total > 0 else 0
            return {"next_name": name, "next_threshold": threshold, "next_remaining": threshold - clan_reputation, "pct": pct, "prev_threshold": prev, "clan_reputation": clan_reputation, "prev_name": prev_name}
        prev = threshold
        prev_name = name
    return None

def load_changes():
    try:
        with open(CHANGES_JSON) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def save_changes(changes):
    now = datetime.now(TARGET_TZ)
    cutoff = now - timedelta(hours=72)
    pruned = [c for c in changes if c.get("detected_at", "") >= cutoff.strftime("%Y-%m-%d %H:%M:%S")]
    with open(CHANGES_JSON, "w") as f:
        json.dump(pruned, f)

def diff_html(diff_str):
    if diff_str.startswith("+"):
        return f'<span class="up">{diff_str}</span>'
    elif diff_str == "0":
        return f'<span class="down">{diff_str}</span>'
    elif diff_str.startswith("-"):
        return f'<span class="down">{diff_str}</span>'
    else:
        return f'<span class="na">{diff_str}</span>'

def save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, all_dates, show_changes, season_info=None, stats=None, diff_30m=None, goal_info=None, changes=None, hourly_cache=None, cache_30m=None):
    daily_rows = compute_diff(data["members"], prev_data)
    clan_name = data.get("clan_name", "Unknown")
    date_str = now.strftime("%Y-%m-%d")
    ts_str = now.strftime("%Y-%m-%d %H:%M:%S")
    member_count = len(data["members"])

    logo_b64 = ""
    logo_path = "clan_logo.png"
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            logo_b64 = base64.b64encode(f.read()).decode()

    favicon_b64 = ""
    if os.path.exists("favicon.ico"):
        with open("favicon.ico", "rb") as f:
            favicon_b64 = base64.b64encode(f.read()).decode()

    archive_links = "".join(
        f'<a href="{d}.html" class="{"active" if d == date_str else ""}">{d}</a>'
        for d in sorted(all_dates, reverse=True)
    )

    diffs_30m_map = diff_30m.get("diffs", {}) if isinstance(diff_30m, dict) else {}
    daily_lookup = {name: diff for name, _, diff in daily_rows}
    uniq_names = get_unique_names(data["members"])
    table_rows = "".join(
        f"<tr><td class=\"num\">{i+1}</td><td>{uniq_names[i][1]}</td><td class=\"num\">{m['member_reputation']}</td><td class=\"num\">{diff_html(diffs_30m_map.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num\">{diff_html(hourly_diffs.get(uniq_names[i][1], 'N/A'))}</td><td class=\"num\">{diff_html(daily_lookup.get(m['character_name'], 'N/A'))}</td></tr>"
        for i, m in enumerate(data["members"])
    )

    changes_html = ""
    if changes:
        left_names = [c["name"] for c in changes if c["type"] == "left"]
        joined_names = [c["name"] for c in changes if c["type"] == "joined"]
        if left_names or joined_names:
            left_items = "".join(f"<li>{n}</li>" for n in left_names)
            joined_items = "".join(f"<li>{n}</li>" for n in joined_names)
            changes_html = f"""
  <div class="changes">
    <div class="changes-title">Member Changes (last 72h)</div>
    <div class="changes-cols">
      <div class="changes-col">
        <div class="changes-head left">Left ({len(left_names)})</div>
        <ul>{left_items}</ul>
      </div>
      <div class="changes-col">
        <div class="changes-head joined">Joined ({len(joined_names)})</div>
        <ul>{joined_items}</ul>
      </div>
    </div>
  </div>"""

    logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo" alt="Clairvoyant">' if logo_b64 else ""
    favicon_html = f'<link rel="icon" type="image/x-icon" href="data:image/x-icon;base64,{favicon_b64}">' if favicon_b64 else ""

    hourly_ref = f"Ref (hourly): {hourly_ts}" if hourly_ts else ""
    ref_30m_val = diff_30m.get("ts", "") if isinstance(diff_30m, dict) else ""
    ref_30m = f"Ref (30m): {ref_30m_val}" if ref_30m_val else ""
    daily_ref = f"Ref (daily): {prev_timestamp}" if prev_timestamp else ""

    timer_html = ""
    season_end_iso = ""
    if season_info:
        season_num = season_info["season"]
        end_dt = datetime(2026, 7, 19, 5, 0, 0, tzinfo=timezone.utc)
        season_end_iso = "2026-07-19T05:00:00Z"
        timer_html = f"""
  <div class="timer-bar">
    <span class="timer-left">
      <span class="timer-season">Season <span id="season-num">{season_num}</span></span>
      <span class="timer-sep">&middot;</span>
      <span class="timer-clock">
        <span class="timer-digits"><span id="timer-d">--</span><span class="timer-unit">d</span></span>
        <span class="timer-digits"><span id="timer-h">--</span><span class="timer-unit">h</span></span>
        <span class="timer-digits"><span id="timer-m">--</span><span class="timer-unit">m</span></span>
        <span class="timer-digits"><span id="timer-s">--</span><span class="timer-unit">s</span></span>
      </span>
    </span>
    <span class="timer-right" id="auto-refresh">&#x21BB; Auto Refresh in <span id="auto-seconds">60</span>s</span>
  </div>"""

    stats_html = ""
    if stats:
        proj_cols = ""
        if "projection" in stats:
            proj_cols = f"""
      <div class="stats-col">
        <span class="stat-label">Est. Season Total</span>
        <span class="stat-val" id="est-season">{stats['projection']:,}</span>
      </div>
      <div class="stats-col">
        <span class="stat-label" id="avg-label">Avg/Day &middot; {stats['days_left']}d left</span>
        <span class="stat-val" id="avg-daily">{stats['avg_daily']:,}</span>
      </div>"""
        stats_html = f"""
  <div class="stats-bar">
    <div class="stats-row">
      <div class="stats-col">
        <span class="stat-label">Today</span>
        <span class="stat-val" id="today-gain">+{stats['today_gain']:,}</span>
      </div>
      <div class="stats-col">
        <span class="stat-label">Season Total</span>
        <span class="stat-val" id="season-total">{stats['season_total']:,}</span>
      </div>{proj_cols}
    </div>
  </div>"""

    goal_html = ""
    if goal_info:
        pct = goal_info["pct"]
        goal_html = f"""
  <div class="goal-bar">
    <div class="goal-track"><div class="goal-fill" style="width:{pct:.1f}%"></div></div>
    <div class="goal-info">
      <span>Next: <span class="goal-next">{goal_info['next_name']}</span> &middot; {pct:.1f}%</span>
      <span class="goal-num">{goal_info['clan_reputation']:,} / {goal_info['next_threshold']:,}</span>
    </div>
  </div>"""

    script_html = ""
    if season_info:
        script_html = """<script>
(function() {
  var end = new Date(\"""" + season_end_iso + """\").getTime();
  function tick() {
    var diff = end - new Date().getTime();
    if (diff <= 0) { document.getElementById("timer-d").textContent = "0"; document.getElementById("timer-h").textContent = "00"; document.getElementById("timer-m").textContent = "00"; document.getElementById("timer-s").textContent = "00"; return; }
    document.getElementById("timer-d").textContent = Math.floor(diff / 86400000);
    document.getElementById("timer-h").textContent = String(Math.floor((diff % 86400000) / 3600000)).padStart(2,"0");
    document.getElementById("timer-m").textContent = String(Math.floor((diff % 3600000) / 60000)).padStart(2,"0");
    document.getElementById("timer-s").textContent = String(Math.floor((diff % 60000) / 1000)).padStart(2,"0");
  }
  tick();
  setInterval(tick, 1000);
})();
window.__goalTiers = [[100000,"5 Stamina Rolls"],[500000,"20 Stamina Rolls"],[750000,"Back Item"],[1000000,"Weapon"],[1600000,"Jutsu"]];
window.__clanId = """ + str(CLAN_ID) + """;
window.__avgDaily = """ + (str(stats['avg_daily']) if stats and 'avg_daily' in stats else '0') + """;
window.__seasonEnd = \"""" + season_end_iso + """\";
window.__hourlyCache = """ + json.dumps(hourly_cache if hourly_cache else {}) + """;
window.__30mCache = """ + json.dumps(cache_30m["members"] if cache_30m and "members" in cache_30m else {}) + """;
(function() {
  var tbody = document.querySelector("tbody");
  window.__originalRows = tbody.innerHTML;
  window.__defaultRows = tbody.innerHTML;
  var sortCol = -1, sortDir = 0;
  var ths = document.querySelectorAll("th");
  function applySort() {
    if (sortDir === 0) { tbody.innerHTML = window.__originalRows; for (var a = 0; a < ths.length; a++) ths[a].querySelector(".sort-arrow").textContent = ""; if (window.__refreshData) window.__refreshData(); var se = document.getElementById("search-input"); if(se&&se.value){var q=se.value.toLowerCase(),rr=tbody.querySelectorAll("tr");for(var ri=0;ri<rr.length;ri++)rr[ri].style.display=rr[ri].cells[1].textContent.trim().toLowerCase().indexOf(q)>=0?"":"none";} return; }
    for (var a = 0; a < ths.length; a++) ths[a].querySelector(".sort-arrow").textContent = "";
    ths[sortCol].querySelector(".sort-arrow").textContent = sortDir === 1 ? "\\u25B2" : "\\u25BC";
    var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
    rows.sort(function(a, b) {
      var va = a.cells[sortCol].textContent.trim(), vb = b.cells[sortCol].textContent.trim();
      if (sortCol === 1) return sortDir === 1 ? va.localeCompare(vb) : vb.localeCompare(va);
      var na = parseFloat(va) || -1/0, nb = parseFloat(vb) || -1/0;
      return sortDir === 1 ? na - nb : nb - na;
    });
    for (var r = 0; r < rows.length; r++) tbody.appendChild(rows[r]);
    var sr = tbody.querySelectorAll("tr");
    for (var ri = 0; ri < sr.length; ri++) sr[ri].cells[0].textContent = ri + 1;
  }
  window.__resetSort = function() { sortCol = -1; sortDir = 0; applySort(); };
  for (var i = 0; i < ths.length; i++) (function(col) {
    ths[col].addEventListener("click", function() {
      if (sortCol !== col) { sortCol = col; sortDir = 1; }
      else { sortDir = (sortDir + 1) % 3; }
      applySort();
      localStorage.setItem("nr_sort", JSON.stringify({col: sortCol, dir: sortDir}));
    });
  })(i);
  try { var _s = JSON.parse(localStorage.getItem("nr_sort")); if (_s && _s.dir > 0) { sortCol = _s.col; sortDir = _s.dir; applySort(); } } catch(e) {}
})();
(function() {
  var API = "https://playninjarift.com/api/detail_clan_website.php?clan_id=" + window.__clanId, RK = "https://playninjarift.com/api/clan_ranking_website.php";
  var tb = document.querySelector("tbody"), names = [], rws = tb.querySelectorAll("tr");
  for (var i = 0; i < rws.length; i++) names.push(rws[i].cells[1].textContent.trim());
  var autoSeconds = 60, autoEl = document.getElementById("auto-seconds"), searchEl = document.getElementById("search-input"), dotEl = document.getElementById("status-dot"), statusEl = document.getElementById("status-text");
  if (window.__hourlyCache && Object.keys(window.__hourlyCache).length > 0) { var _ts = ts(), _m = String(new Date().getMinutes()), _b = _m <= "1" ? "01" : (_m >= "31" && _m <= "32" ? "31" : _m); localStorage.setItem("nr_1h", JSON.stringify({b: _b, ts: _ts, rs: window.__hourlyCache})); }
  if (window.__30mCache && Object.keys(window.__30mCache).length > 0) { var _b30 = (new Date().getMinutes() <= 1 ? "01" : "31"); localStorage.setItem("nr_30m", JSON.stringify({b: _b30, ts: _ts, rs: window.__30mCache})); }
  function pad(n) { return n < 10 ? "0"+n : ""+n; }
  function ts() { var d = new Date(); return d.getFullYear()+"-"+pad(d.getMonth()+1)+"-"+pad(d.getDate())+" "+pad(d.getHours())+":"+pad(d.getMinutes())+":"+pad(d.getSeconds()); }
  function fj(u) { return fetch(u,{headers:{"Accept":"application/json"}}).then(function(r){return r.json();}).catch(function(){return null;}); }
  function dh(v) { return v > 0 ? '<span class="up">+'+v+"</span>" : v <= 0 ? '<span class="down">'+v+"</span>" : '<span class="na">N/A</span>'; }
  function blk30(m) { return m <= 1 ? "01" : (m >= 31 && m <= 32 ? "31" : null); }
  function blk1h(m) { return m <= 1 ? "01" : null; }
  function upd(d, rk) {
    autoSeconds = 60;
    var clan = null;
    if (rk) {
      if (Array.isArray(rk)) {
        for (var ci = 0; ci < rk.length; ci++) {
          if (rk[ci].clan_id === window.__clanId) { clan = rk[ci]; break; }
        }
      } else { clan = rk; }
    }
    var n = new Date(), nm = n.getMinutes(), ns = ts();
    var lm = {}; for (var i = 0; i < d.length; i++) lm[i < names.length ? names[i] : d[i].character_name] = d[i].member_reputation;
    var n2r = {}, a = tb.querySelectorAll("tr"); for (var i = 0; i < a.length; i++) n2r[a[i].cells[1].textContent.trim()] = a[i];
    var c30 = null, c1h = null; try { c30 = JSON.parse(localStorage.getItem("nr_30m")); c1h = JSON.parse(localStorage.getItem("nr_1h")); } catch(e) {}
    if (!c1h && window.__hourlyCache) c1h = {rs: window.__hourlyCache, ts: ""};
    for (var i = 0; i < names.length; i++) {
      var name = names[i], rep = lm[name]; if (rep === undefined) continue;
      var row = n2r[name]; if (!row) continue;
      var cel = row.cells;
      cel[2].textContent = rep;
      if (c30 && c30.rs && c30.rs[name] !== undefined) cel[3].innerHTML = dh(rep - c30.rs[name]);
      if (c1h && c1h.rs && c1h.rs[name] !== undefined) cel[4].innerHTML = dh(rep - c1h.rs[name]);
    }
    for (var _n in lm) {
      if (names.indexOf(_n) === -1) {
        names.push(_n);
        var tr = document.createElement("tr");
        tr.className = "new-row";
        tr.innerHTML = '<td class="num"></td><td>' + _n + '</td><td class="num">' + lm[_n] + '</td><td class="num"><span class="na">N/A</span></td><td class="num"><span class="na">N/A</span></td><td class="num"><span class="na">N/A</span></td>';
        tb.appendChild(tr);
        if (searchEl && searchEl.value && _n.toLowerCase().indexOf(searchEl.value.toLowerCase()) === -1) tr.style.display = "none";
      }
    }
    a = tb.querySelectorAll("tr"); n2r = {};
    for (var i = 0; i < a.length; i++) n2r[a[i].cells[1].textContent.trim()] = a[i];
    for (var i = 0; i < names.length; i++) {
      var row = n2r[names[i]];
      if (row && lm[names[i]] === undefined) row.className = "left-row";
    }
    if (clan) {
      var te = document.getElementById("today-gain");
      if (te && clan.clan_day_points !== undefined) te.textContent = "+"+Number(clan.clan_day_points).toLocaleString();
      var sv = document.querySelectorAll(".stats-col .stat-val");
      if (sv.length >= 2 && clan.clan_reputation !== undefined) sv[1].textContent = Number(clan.clan_reputation).toLocaleString();
    }
    var ft = document.querySelector(".footer");
    var st = document.getElementById("snapshot-ts"); if (st) st.textContent = ns;
    var rs = {}; for (var i = 0; i < d.length; i++) rs[d[i].character_name] = d[i].member_reputation;
    var b30 = blk30(nm), b1h = blk1h(nm);
    if (b30 && (!c30 || c30.b !== b30)) localStorage.setItem("nr_30m", JSON.stringify({b: b30, ts: ns, rs: rs}));
    if (b1h && (!c1h || c1h.b !== b1h)) localStorage.setItem("nr_1h", JSON.stringify({b: b1h, ts: ns, rs: rs}));
    window.__defaultRows = tb.innerHTML;
    var sr = tb.querySelectorAll("tr");
    for (var ri = 0; ri < sr.length; ri++) sr[ri].cells[0].textContent = ri + 1;
    if (clan && clan.clan_reputation !== undefined) {
      var rep = Number(clan.clan_reputation), prev = 0, tiers = window.__goalTiers;
      for (var ti = 0; ti < tiers.length; ti++) {
        if (rep < tiers[ti][0]) {
          var pct = ((rep - prev) / (tiers[ti][0] - prev)) * 100;
          var fel = document.querySelector(".goal-fill");
          var nel = document.querySelector(".goal-info .goal-num");
          var nel2 = document.querySelector(".goal-info .goal-next");
          if (fel) fel.style.width = pct.toFixed(1) + "%";
          if (nel) nel.textContent = rep.toLocaleString() + " / " + tiers[ti][0].toLocaleString();
          if (nel2) nel2.textContent = tiers[ti][1];
          break;
        }
        prev = tiers[ti][0];
      }
    }
    if (clan && window.__avgDaily > 0 && window.__seasonEnd) {
      var seEnd = new Date(window.__seasonEnd).getTime(), nowMs = new Date().getTime();
      var daysLeft = Math.ceil((seEnd - nowMs) / 86400000);
      if (daysLeft < 0) daysLeft = 0;
      var proj = Number(clan.clan_reputation);
      var cur = new Date(new Date().getFullYear(), new Date().getMonth(), new Date().getDate() + 1, 0, 0, 0, 0);
      while (cur.getTime() <= seEnd) {
        proj += cur.getDay() % 6 === 0 ? window.__avgDaily * 2 : window.__avgDaily;
        cur = new Date(cur.getTime() + 86400000);
      }
      var estEl = document.getElementById("est-season");
      var avgLabel = document.getElementById("avg-label");
      if (estEl) estEl.textContent = Math.round(proj).toLocaleString();
      if (avgLabel) avgLabel.textContent = "Avg/Day · " + daysLeft + "d left";
    }
  }
  function refreshData() {
    if (dotEl) dotEl.className = "status-dot wait";
    if (statusEl) statusEl.textContent = "Loading...";
    Promise.all([fj(API), fj(RK)]).then(function(r) {
      if (r[0] && r[0].members) {
        upd(r[0].members, r[1]);
        if (dotEl) dotEl.className = "status-dot ok";
        if (statusEl) statusEl.textContent = "Live";
      } else {
        if (dotEl) dotEl.className = "status-dot err";
        if (statusEl) statusEl.textContent = "Offline";
      }
    });
  }
  window.__refreshData = refreshData;
  refreshData();
  setInterval(refreshData, 60000);
  setInterval(function(){if(autoSeconds>0)autoSeconds--;if(autoEl)autoEl.textContent=autoSeconds;},1000);
  if(autoEl)autoEl.parentElement.addEventListener("click",function(){autoSeconds=60;refreshData();});
  if(searchEl)searchEl.addEventListener("input",function(){
    var q = this.value.toLowerCase(), r = tb.querySelectorAll("tr");
    for(var i=0;i<r.length;i++)r[i].style.display=r[i].cells[1].textContent.trim().toLowerCase().indexOf(q)>=0?"":"none";
  });
  // Click-to-copy (event delegation on tbody)
  tb.addEventListener("click", function(e) {
    var cell = e.target;
    while (cell && cell.tagName !== "TD") cell = cell.parentNode;
    if (!cell || cell.cellIndex !== 1) return;
    var name = cell.textContent.trim();
    navigator.clipboard.writeText(name).then(function() {
      var toast = document.createElement("span");
      toast.className = "copied-toast";
      toast.textContent = "Copied!";
      toast.style.opacity = "1";
      cell.style.position = "relative";
      cell.appendChild(toast);
      setTimeout(function() { toast.style.opacity = "0"; setTimeout(function() { toast.remove(); }, 300); }, 1500);
    }).catch(function() {});
  });
  // Reset sort
  var resetBtn = document.getElementById("reset-btn");
  if (resetBtn) resetBtn.addEventListener("click", function() {
    localStorage.removeItem("nr_sort");
    if (window.__resetSort) window.__resetSort();
  });
  // CSV export
  function csvDownload() {
    var rows = tb.querySelectorAll("tr"), csv = "Rank,Name,Total Reps,1/2 Hour,Hourly,Daily\\n";
    for (var i = 0; i < rows.length; i++) {
      var cells = rows[i].cells, vals = [];
      for (var j = 0; j < cells.length; j++) {
        var v = cells[j].textContent.trim().replace(/"/g, '""');
        vals.push('"' + v + '"');
      }
      csv += vals.join(",") + "\\n";
    }
    var blob = new Blob([csv], {type: "text/csv;charset=utf-8"});
    var url = URL.createObjectURL(blob);
    var a = document.createElement("a");
    a.href = url; a.download = "clairvoyant_reps.csv"; a.click();
    URL.revokeObjectURL(url);
  }
  var csvLink = document.getElementById("csv-link");
  if (csvLink) csvLink.addEventListener("click", csvDownload);
  // Updated ago
  function updateAgo() {
    var st = document.getElementById("snapshot-ts");
    var ua = document.getElementById("updated-ago");
    if (!st || !ua) return;
    var txt = st.textContent;
    var m = txt.match(/(\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2})/);
    if (!m) return;
    var snap = new Date(m[1].replace(" ", "T") + "+08:00").getTime();
    var now = new Date().getTime();
    var diff = Math.floor((now - snap) / 60000);
    ua.textContent = "Updated " + (diff < 1 ? "just now" : diff + "m ago");
  }
  updateAgo();
  setInterval(updateAgo, 30000);
})();
</script>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{favicon_html}
<title>{clan_name} [Reps]</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
    background: #080810;
    color: #e0e0e0;
    min-height: 100vh;
    display: flex;
    justify-content: center;
    padding: 32px 16px;
  }}
  .container {{
    max-width: 960px;
    width: 100%;
    box-shadow: 0 0 40px rgba(233, 69, 96, 0.06), 0 8px 32px rgba(0,0,0,0.5);
    border-radius: 16px;
    overflow: hidden;
  }}
  .header {{
    text-align: center;
    padding: 32px 24px 24px;
    background: linear-gradient(135deg, #0f0f1e 0%, #1a1a30 50%, #0d1b2a 100%);
    position: relative;
    overflow: hidden;
  }}
  .header::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #e94560, #ff6b8a, #e94560);
    background-size: 200% 100%;
    animation: shimmer 3s ease-in-out infinite;
  }}
  @keyframes shimmer {{ 0%,100% {{ background-position: 0% 50%; }} 50% {{ background-position: 100% 50%; }} }}
  .logo {{
    width: 132px; height: 132px;
    object-fit: contain;
    margin-bottom: 16px;
    filter: drop-shadow(0 0 16px rgba(233, 69, 96, 0.3));
  }}
  .header h1 {{
    font-size: 30px;
    font-weight: 700;
    color: #fff;
    letter-spacing: 0.5px;
    margin-bottom: 4px;
  }}
  .header .sub {{
    font-size: 17px;
    color: #888;
    display: flex;
    justify-content: center;
    gap: 16px;
    flex-wrap: wrap;
  }}
  .header .sub span {{ color: #aaa; }}
  .archive {{
    display: flex;
    gap: 6px;
    justify-content: center;
    padding: 14px 20px;
    background: #0c0c18;
    border-bottom: 1px solid #1a1a2e;
    border-top: 1px solid #1a1a2e;
  }}
  .archive a {{
    color: #777;
    text-decoration: none;
    font-size: 12px;
    padding: 5px 14px;
    border-radius: 20px;
    border: 1px solid #1a1a2e;
    transition: 0.25s;
  }}
  .archive a:hover {{ border-color: #e94560; color: #fff; background: rgba(233, 69, 96, 0.08); }}
  .archive a.active {{ border-color: #e94560; color: #fff; background: #e94560; font-weight: 600; }}
  .table-wrap {{ overflow-x: auto; }}
  table {{
    width: 100%;
    min-width: 0;
    border-collapse: collapse;
    background: #0c0c14;
  }}
  thead {{ position: sticky; top: 0; z-index: 1; }}
  th {{
    background: #0f0f1e;
    padding: 14px 18px;
    text-align: center;
    font-size: 15px;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #e94560;
    font-weight: 600;
  }}
  td {{
    padding: 11px 18px;
    border-bottom: 1px solid #14141f;
    font-size: 14px;
    color: #ccc;
    text-align: center;
  }}
  tr:nth-child(even) td {{ background: rgba(255,255,255,0.015); }}
  tr:hover td {{ background: rgba(233, 69, 96, 0.04); }}
  td.num {{ font-variant-numeric: tabular-nums; }}
  .up {{ color: #4caf50; }}
  .down {{ color: #f44336; }}
  .na {{ color: #555; }}
  .changes {{
    background: #0c0c14;
    padding: 20px 24px;
    border-top: 1px solid #14141f;
  }}
  .changes-title {{
    font-size: 13px;
    color: #e94560;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 14px;
    text-align: center;
  }}
  .changes-cols {{
    display: flex;
    gap: 24px;
    justify-content: center;
  }}
  .changes-col {{
    flex: 1;
    max-width: 320px;
  }}
  .changes-head {{
    font-size: 12px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-weight: 600;
    margin-bottom: 8px;
    text-align: center;
  }}
  .changes-head.left {{ color: #f44336; }}
  .changes-head.joined {{ color: #4caf50; }}
  .changes-col ul {{
    list-style: none;
    padding: 0;
    margin: 0;
    text-align: center;
  }}
  .changes-col li {{
    padding: 4px 0;
    font-size: 14px;
    color: #ccc;
    border-bottom: 1px solid #14141f;
  }}
  .changes-col li:last-child {{ border-bottom: none; }}
  .footer {{
    text-align: center;
    padding: 18px 20px;
    background: #08080f;
    color: #444;
    font-size: 12px;
    border-top: 1px solid #12121e;
  }}
  .footer .ref {{ color: #555; font-size: 11px; margin-top: 2px; }}
  .footer a {{ color: #e94560; text-decoration: none; }}
  .timer-bar {{
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 10px;
    padding: 12px 20px;
    background: #0f142373;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
    font-size: 15px;
    flex-wrap: wrap;
    position: relative;
  }}
  .timer-left {{ display: flex; align-items: center; gap: 10px; }}
  .timer-season {{ color: #eab308; font-weight: 700; letter-spacing: 0.5px; }}
  .timer-sep {{ color: #444; }}
  .timer-clock {{ display: flex; align-items: center; gap: 6px; }}
  .timer-digits {{ font-variant-numeric: tabular-nums; }}
  .timer-digits span:first-child {{ color: #2dd4bf; font-weight: 600; min-width: 28px; display: inline-block; text-align: center; }}
  .timer-unit {{ color: #888; font-size: 12px; margin-left: 1px; }}
  .timer-right {{ position: absolute; right: 20px; top: 50%; transform: translateY(-50%); cursor: pointer; font-size: 12px; color: #888; user-select: none; white-space: nowrap; }}
  .timer-right:hover {{ color: #e94560; }}
  .stats-bar {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 8px;
    padding: 14px 20px;
    background: #0f142373;
    backdrop-filter: blur(8px);
    -webkit-backdrop-filter: blur(8px);
    border-top: 1px solid #1a1a2e;
  }}
  .stats-row {{
    display: flex;
    justify-content: center;
    gap: 48px;
    width: 100%;
  }}
  .stats-col {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 2px;
  }}
  .stat-label {{ color: #888; font-size: 12px; text-transform: uppercase; letter-spacing: 0.3px; }}
  .stat-val {{ color: #e0e0e0; font-size: 18px; font-weight: 700; font-variant-numeric: tabular-nums; }}
  #today-gain {{ color: #4caf50; }}
  th {{ cursor: pointer; user-select: none; }}
  th .sort-arrow {{ font-size: 11px; margin-left: 3px; }}
  @media (max-width: 600px) {{
    body {{ padding: 16px 8px; }}
    .header {{ padding: 24px 16px 20px; }}
    .header h1 {{ font-size: 22px; }}
    .header .sub {{ font-size: 14px; }}
    .logo {{ width: 96px; height: 96px; }}
    table {{ min-width: 520px; }}
    th, td {{ padding: 10px 10px; font-size: 12px; }}
    .changes {{ padding: 16px; }}
    .changes-cols {{ flex-direction: column; gap: 14px; }}
    .changes-col {{ max-width: 100%; }}
    .archive {{ flex-wrap: nowrap; overflow-x: auto; justify-content: flex-start; -webkit-overflow-scrolling: touch; scrollbar-width: none; }}
    .archive::-webkit-scrollbar {{ display: none; }}
    .archive a {{ flex-shrink: 0; }}
    .stats-bar {{ gap: 10px; padding: 12px 16px; }}
    .stat-val {{ font-size: 15px; }}
    .stats-row {{ flex-direction: column; gap: 6px; align-items: center; }}
    .stats-row + .stats-row {{ border-top: 1px solid #1a1a2e; padding-top: 10px; }}
  }}
  .live-bar {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 8px 20px;
    background: #0a0a14;
    border-top: 1px solid #14141f;
    border-bottom: 1px solid #14141f;
    gap: 12px;
    flex-wrap: wrap;
  }}
  #search-input {{
    flex: 1;
    min-width: 160px;
    padding: 7px 12px;
    border-radius: 6px;
    border: 1px solid #1a1a2e;
    background: #0f0f1e;
    color: #e0e0e0;
    font-size: 13px;
    outline: none;
  }}
  #search-input:focus {{ border-color: #e94560; }}
  #search-input::placeholder {{ color: #555; }}
  .live-status {{ display: flex; align-items: center; gap: 6px; font-size: 12px; color: #555; white-space: nowrap; }}
  .status-dot {{ width: 8px; height: 8px; border-radius: 50%; }}
  .status-dot.ok {{ background: #4caf50; }}
  .status-dot.err {{ background: #f44336; }}
  .status-dot.wait {{ background: #888; animation: pulse 1.5s infinite; }}
  @keyframes pulse {{ 0%,100% {{ opacity: 1; }} 50% {{ opacity: 0.3; }} }}
  tr.new-row td {{ animation: fadeIn 0.5s ease; }}
  @keyframes fadeIn {{ from {{ opacity: 0; background: rgba(233,69,96,0.1); }} to {{ opacity: 1; background: transparent; }} }}
  tr.left-row td {{ opacity: 0.35; }}
  .goal-bar {{
    display: flex;
    flex-direction: column;
    gap: 4px;
    padding: 14px 20px;
    background: #0c0c18;
    border-top: 1px solid #1a1a2e;
  }}
  .goal-track {{
    width: 100%;
    height: 16px;
    background: #14141f;
    border-radius: 8px;
    overflow: hidden;
  }}
  .goal-fill {{
    height: 100%;
    background: linear-gradient(90deg, #e94560, #ff6b8a);
    border-radius: 8px;
    transition: width 0.5s ease;
  }}
  .goal-info {{
    display: flex;
    justify-content: space-between;
    font-size: 12px;
    color: #888;
  }}
  .goal-info .goal-next {{ color: #e94560; font-weight: 600; }}
  .goal-info .goal-num {{ color: #ccc; font-variant-numeric: tabular-nums; }}
  td:first-child, th:first-child {{ width: 28px; min-width: 28px; text-align: center; color: #666; font-size: 12px; }}
  .action-btn {{ cursor: pointer; font-size: 12px; color: #888; padding: 4px 10px; border-radius: 4px; border: 1px solid #1a1a2e; background: #0f0f1e; user-select: none; white-space: nowrap; }}
  .action-btn:hover {{ border-color: #e94560; color: #e94560; }}
  .footer-updated {{ color: #555; font-size: 11px; margin: 2px 0; }}
  .footer-csv {{ margin-top: 8px; }}
  .footer-csv a {{ color: #e94560; text-decoration: none; font-size: 12px; cursor: pointer; }}
  .footer-csv a:hover {{ text-decoration: underline; }}
  .copied-toast {{ position: absolute; background: #e94560; color: #fff; font-size: 11px; padding: 2px 8px; border-radius: 4px; white-space: nowrap; pointer-events: none; opacity: 0; transition: opacity 0.3s; z-index: 10; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    {logo_html}
    <h1>{clan_name}</h1>
    <div class="sub">
      <span>Clan ID: {CLAN_ID}</span>
      <span>&middot;</span>
      <span>{member_count} members</span>
    </div>
  </div>
  {timer_html}
  {stats_html}
  {goal_html}
  {f'<div class="archive">{archive_links}</div>' if archive_links else ""}
  <div class="live-bar">
    <div style="display:flex;align-items:center;gap:8px;flex:1;flex-wrap:wrap">
      <input type="text" id="search-input" placeholder="Search member...">
      <span class="action-btn" id="reset-btn">Reset</span>
    </div>
    <div class="live-status">
      <span class="status-dot wait" id="status-dot"></span>
      <span id="status-text">Idle</span>
    </div>
  </div>
  <div class="table-wrap">
  <table>
    <thead><tr><th># <span class="sort-arrow"></span></th><th>Name <span class="sort-arrow"></span></th><th>Total Reps <span class="sort-arrow"></span></th><th>1/2 Hour <span class="sort-arrow"></span></th><th>Hourly <span class="sort-arrow"></span></th><th>Daily <span class="sort-arrow"></span></th></tr></thead>
    <tbody>{table_rows}</tbody>
  </table>
  </div>
  {changes_html}
  <div class="footer">
    <div class="footer-updated" id="updated-ago"></div>
    Snapshot: <span id="snapshot-ts">{ts_str}</span>
    <div class="ref">{ref_30m}{" &middot; " if ref_30m and (hourly_ref or daily_ref) else ""}{hourly_ref}{" &middot; " if hourly_ref and daily_ref else ""}{daily_ref}</div>
    <div class="footer-csv"><a id="csv-link">Download CSV</a></div>
  </div>
</div>
{script_html}
</body>
</html>"""

    index_path = "index.html"
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[{ts_str}] Saved {index_path}")

    if all_dates:
        archive_path = f"{date_str}.html"
        with open(archive_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"[{ts_str}] Saved {archive_path}")

def save_snapshot(data):
    now = datetime.now(TARGET_TZ)
    is_daily = (now.hour == 13)
    sheet_name = now.strftime("%Y-%m-%d")

    prev_data, prev_timestamp = load_prev_from_xlsx(EXCEL_FILE, sheet_name)

    uniq = get_unique_names(data["members"])

    cache_1h = load_1h_cache()
    hourly_diffs = {}
    hourly_ts = ""
    if cache_1h and cache_1h.get("members"):
        hourly_ts = cache_1h.get("timestamp", "")
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            reps = m["member_reputation"]
            memb = cache_1h["members"]
            if uni_name in memb:
                diff = reps - memb[uni_name]
                hourly_diffs[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                hourly_diffs[uni_name] = "N/A"

    cache_30m = load_30m_cache()
    diffs_30m = {}
    ref_30m_ts = ""
    if cache_30m:
        ref_30m_ts = cache_30m.get("timestamp", "")
        for i, m in enumerate(data["members"]):
            uni_name = uniq[i][1]
            reps = m["member_reputation"]
            if uni_name in cache_30m.get("members", {}):
                diff = reps - cache_30m["members"][uni_name]
                diffs_30m[uni_name] = f"+{diff}" if diff > 0 else str(diff)
            else:
                diffs_30m[uni_name] = "N/A"
    diff_30m_data = {"ts": ref_30m_ts, "diffs": diffs_30m}

    try:
        season_info = fetch_season_info()
    except Exception:
        season_info = None

    try:
        ranking = fetch_clan_ranking()
        clan_reputation = ranking.get("clan_reputation", 0)
        today_gain = ranking.get("clan_day_points", 0)
    except Exception:
        clan_reputation = 0
        today_gain = 0

    changes = load_changes()
    if cache_30m and "order" in cache_30m:
        now_ts = now.strftime("%Y-%m-%d %H:%M:%S")
        raw_prev = [re.sub(r' \(#\d+\)$', '', n) for n in cache_30m["order"]]
        raw_curr = [m["character_name"] for m in data["members"]]
        from collections import Counter
        prev_count, curr_count = Counter(raw_prev), Counter(raw_curr)
        existing_change_keys = set((c["type"], c["name"]) for c in changes)
        for name, cnt in prev_count.items():
            diff = cnt - curr_count.get(name, 0)
            for _ in range(diff):
                if ("left", name) not in existing_change_keys:
                    changes.append({"type": "left", "name": name, "detected_at": now_ts})
        for name, cnt in curr_count.items():
            diff = cnt - prev_count.get(name, 0)
            for _ in range(diff):
                if ("joined", name) not in existing_change_keys:
                    changes.append({"type": "joined", "name": name, "detected_at": now_ts})
        save_changes(changes)
 
    goal_info = compute_goal_info(clan_reputation)
    stats = None
    if season_info:
        end_dt = datetime(2026, 7, 19, 5, 0, 0, tzinfo=timezone.utc)
        avg_daily = compute_rolling_avg_daily_gain(EXCEL_FILE, sheet_name)
        proj = compute_season_projection(clan_reputation, avg_daily, end_dt, now)
        stats = {"today_gain": today_gain, "season_total": clan_reputation}
        if proj:
            stats["projection"] = proj["projection"]
            stats["avg_daily"] = proj["avg_daily"]
            stats["days_left"] = proj["days_left"]
 
    is_hourly_mark = (now.minute <= 1)

    if is_daily:
        save_xlsx(data, prev_data, now, uniq)
        existing_html = [f.replace(".html", "") for f in os.listdir(".") if f.endswith(".html") and f[:4].isdigit() and f != "index.html"]
        all_dates = set(existing_html)
        all_dates.add(sheet_name)
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, sorted(all_dates), show_changes=True, season_info=season_info, stats=stats, diff_30m=diff_30m_data, goal_info=goal_info, changes=changes, hourly_cache=cache_1h["members"] if cache_1h else {}, cache_30m=cache_30m)
    else:
        save_html(data, prev_data, prev_timestamp, hourly_diffs, hourly_ts, now, [], show_changes=False, season_info=season_info, stats=stats, diff_30m=diff_30m_data, goal_info=goal_info, changes=changes, hourly_cache=cache_1h["members"] if cache_1h else {}, cache_30m=cache_30m)

    save_30m_cache(data["members"], uniq)
    if cache_30m:
        save_1h_cache(cache_30m.get("members", {}), cache_30m.get("timestamp", ""))
    if is_hourly_mark:
        save_hourly_cache(data["members"], uniq, now)

    if season_info and now >= end_dt and avg_daily is not None:
        save_seasonal_xlsx(data["members"], season_info["season"])

    try:
        save_daily_history()
    except Exception as e:
        print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] History error: {e}")

def save_daily_history():
    if not os.path.exists(EXCEL_FILE):
        return
    wb = load_workbook(EXCEL_FILE)
    names = sorted([s.title for s in wb.worksheets if s.title != "Sheet1"])
    if len(names) < 2:
        return
    sheets_data = []
    for s in names:
        ws = wb[s]
        members = []
        for row in ws.iter_rows(min_row=4, max_col=2, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            if name and row[1] is not None:
                members.append((name, int(row[1])))
        sheets_data.append({"date": s, "members": members})
    css = """<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'Segoe UI', system-ui, sans-serif; background: #080810; color: #e0e0e0; min-height: 100vh; display: flex; justify-content: center; padding: 32px 16px; }
  .container { max-width: 800px; width: 100%; box-shadow: 0 0 40px rgba(233,69,96,0.06), 0 8px 32px rgba(0,0,0,0.5); border-radius: 16px; overflow: hidden; }
  .header { text-align: center; padding: 28px 24px 20px; background: linear-gradient(135deg, #0f0f1e 0%, #1a1a30 50%, #0d1b2a 100%); }
  .header h1 { font-size: 26px; font-weight: 700; color: #fff; margin-bottom: 4px; }
  .header .sub { font-size: 14px; color: #888; }
  .nav { display: flex; justify-content: space-between; padding: 12px 20px; background: #0c0c18; border-top: 1px solid #1a1a2e; border-bottom: 1px solid #1a1a2e; }
  .nav a { color: #e94560; text-decoration: none; font-size: 13px; }
  .nav a:hover { text-decoration: underline; }
  .nav .inactive { color: #444; pointer-events: none; }
  .summary { text-align: center; padding: 10px 20px; background: #0a0a14; color: #888; font-size: 13px; border-bottom: 1px solid #14141f; }
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; background: #0c0c14; }
  th { background: #0f0f1e; padding: 12px 16px; text-align: center; font-size: 13px; text-transform: uppercase; letter-spacing: 1px; color: #e94560; font-weight: 600; }
  td { padding: 10px 16px; border-bottom: 1px solid #14141f; font-size: 13px; color: #ccc; text-align: center; }
  tr:nth-child(even) td { background: rgba(255,255,255,0.015); }
  .green td { color: #4caf50; }
  .red td { color: #f44336; }
  .footer { text-align: center; padding: 16px 20px; background: #08080f; color: #444; font-size: 12px; border-top: 1px solid #12121e; }
  .footer a { color: #e94560; text-decoration: none; }
  .footer a:hover { text-decoration: underline; }
  .index-list { padding: 20px; background: #0c0c14; }
  .index-list a { display: block; padding: 8px 14px; color: #ccc; text-decoration: none; font-size: 14px; border-bottom: 1px solid #14141f; }
  .index-list a:hover { background: rgba(233,69,96,0.04); color: #fff; }
  .index-list a:last-child { border-bottom: none; }
  .index-list .met { color: #4caf50; font-weight: 600; }
  .star-joined { color: #42a5f5; }
  .star-left { color: #f44336; }
</style>"""
    daily_pages = []
    for i in range(1, len(sheets_data)):
        prev, curr = sheets_data[i-1], sheets_data[i]
        date = curr["date"]
        dt = datetime.strptime(date, "%Y-%m-%d")
        day_name = dt.strftime("%a")
        threshold = 1000 if dt.weekday() in (6, 0) else 500
        prev_list, curr_list = prev["members"], curr["members"]
        max_len = max(len(prev_list), len(curr_list))
        gains = []
        for j in range(max_len):
            if j < len(prev_list) and j < len(curr_list):
                pname, prep = prev_list[j]
                cname, crep = curr_list[j]
                gains.append({"name": cname, "gain": crep - prep, "joined": False, "left": False})
            elif j < len(curr_list):
                cname, crep = curr_list[j]
                gains.append({"name": cname, "gain": None, "joined": True, "left": False})
            else:
                pname, prep = prev_list[j]
                gains.append({"name": pname, "gain": None, "joined": False, "left": True})
        gains.sort(key=lambda x: (x["gain"] is None, -(x["gain"] or 0)))
        met_count = sum(1 for g in gains if g["gain"] is not None and g["gain"] >= threshold)
        total_current = len(curr["members"])
        daily_pages.append({"date": date, "day_name": day_name, "threshold": threshold, "met": met_count, "total": total_current})
        rows_html = ""
        for idx, g in enumerate(gains, 1):
            star = ""
            if g["joined"]: star = '<span class="star-joined">&#9733;</span> '
            elif g["left"]: star = '<span class="star-left">&#9734;</span> '
            gain_str = f'+{g["gain"]:,}' if g["gain"] is not None else '<span class="star-left">N/A</span>'
            row_class = ""
            if g["gain"] is not None:
                row_class = "green" if g["gain"] >= threshold else "red"
            rows_html += f'<tr class="{row_class}"><td>{idx}</td><td>{star}{g["name"]}</td><td>{gain_str}</td><td>{"✅" if g["gain"] is not None and g["gain"] >= threshold else "❌"}</td></tr>\n'
        prev_link = f'<a href="history_{prev["date"]}.html">&larr; Previous</a>' if i > 1 else '<span class="inactive">&larr; Previous</span>'
        next_link = f'<a href="history_{sheets_data[i+1]["date"]}.html">Next &rarr;</a>' if i < len(sheets_data) - 1 else '<span class="inactive">Next &rarr;</span>'
        page_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Daily Reps · {date} ({day_name})</title>
{css}
</head>
<body>
<div class="container">
  <div class="header"><h1>Clairvoyant</h1><div class="sub">Daily Reps · {date} ({day_name}) &middot; Min: <b>{threshold:,}</b></div></div>
  <div class="summary">{met_count} of {total_current} members met the threshold ({threshold:,})</div>
  <div class="nav">{prev_link}<a href="history.html">Index</a>{next_link}</div>
  <div class="table-wrap"><table><thead><tr><th>#</th><th>Name</th><th>Gain</th><th>Status</th></tr></thead><tbody>{rows_html}</tbody></table></div>
  <div class="footer"><a href="index.html">&larr; Back to main page</a></div>
</div>
</body>
</html>"""
        with open(f"history_{date}.html", "w", encoding="utf-8") as f:
            f.write(page_html)
        print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] Saved history_{date}.html")
    index_rows = ""
    for dp in daily_pages:
        index_rows += f'<a href="history_{dp["date"]}.html">{dp["date"]} ({dp["day_name"]}) <span class="met">{dp["met"]}/{dp["total"]}</span> met &rarr;</a>\n'
    index_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Daily Rep History · Season 61</title>
{css}
</head>
<body>
<div class="container">
  <div class="header"><h1>Clairvoyant</h1><div class="sub">Daily Rep History (Season 61)</div></div>
  <div class="index-list">{index_rows}</div>
  <div class="footer"><a href="index.html">&larr; Back to main page</a> &middot; <a href="https://github.com/nixervo/Clairvoyant-Reps">Source</a></div>
</div>
</body>
</html>"""
    with open("history.html", "w", encoding="utf-8") as f:
        f.write(index_html)
    print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] Saved history.html")

def run():
    print("Clan snapshot daemon started. Running every 30 minutes.")
    os.system('git config user.name "clan-snapshot-bot"')
    os.system('git config user.email "bot@clan-snapshot.local"')
    while True:
        now = datetime.now(TARGET_TZ)
        if now.minute < 1:
            target = now.replace(minute=1, second=0, microsecond=0)
        elif now.minute < 31:
            target = now.replace(minute=31, second=0, microsecond=0)
        else:
            target = (now + timedelta(hours=1)).replace(minute=1, second=0, microsecond=0)
        sleep_sec = (target - now).total_seconds()
        time.sleep(sleep_sec)
        try:
            data = fetch_clan()
            save_snapshot(data)
            ts = datetime.now(TARGET_TZ).strftime("%Y-%m-%d %H:%M:%S")
            os.system("git add -A")
            os.system(f'git commit -m "auto: snapshot {ts}" --allow-empty')
            os.system("git push")
        except Exception as e:
            print(f"[{datetime.now(TARGET_TZ).strftime('%Y-%m-%d %H:%M:%S')}] ERROR: {e}")

def fetch_once():
    data = fetch_clan()
    save_snapshot(data)

if __name__ == "__main__":
    if "--once" in sys.argv:
        fetch_once()
    else:
        run()
